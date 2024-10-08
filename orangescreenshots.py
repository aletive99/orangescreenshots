import numpy as np
import cv2 as cv
import requests
from bs4 import BeautifulSoup
import os
import urllib
import time
from tqdm import tqdm
import imageio
from datetime import date, timedelta
import yaml
import openpyxl
from openai import OpenAI
from graphlib import TopologicalSorter
import pandas as pd
from transformers import BertModel, BertTokenizer
import torch
import pickle
from sklearn.metrics.pairwise import cosine_similarity


def _get_filenames(direct, ext='image'):
    """
    This function gets the filenames of the widgets from the folder created by the get_widgets function. By default, the
    extension is set to 'image' meaning that only images of the type '.png' and '.jpg' will be extracted, but it can be
    changed according to the needs of the user by specifying an actual extension. The function returns a list with the
    full path of the widgets with the desired extension
    :param direct: str
    :param ext: str
    :return: img_names_tgt: np.array
    """
    img_names_tgt = []
    for root, dirs, filenames in os.walk(direct):
        for filename in filenames:
            if ext == 'image':
                if filename.endswith('.png') or filename.endswith('.jpg') or filename.endswith('.jpeg'):
                    img_names_tgt.append(os.path.join(root, filename))
            elif filename.endswith(ext):
                img_names_tgt.append(os.path.join(root, filename))
    return img_names_tgt


def _download_widgets():
    """
    These operations guarantee that the widgets are downloaded and updated everytime the library is imported.
    """
    main = 'https://orangedatamining.com/'
    sub = 'widget-catalog/'
    url = main + sub
    wd = os.getcwd()
    destination_dir = wd + '/widgets'
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36'}
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        img_tags = soup.find_all('img')
        os.makedirs(destination_dir, exist_ok=True)
        n_iter = len(img_tags)
        i = 0
        remember = -1
        only_check = False
        if len(_get_filenames(destination_dir)) > 200:
            only_check = True
        if not only_check:
            progress_bar = tqdm(total=n_iter, desc="Progress")
            print('Checking and downloading widgets')
        while i < n_iter:
            img = img_tags[i]
            img_url = img.get('src')
            img_url_parsed = urllib.parse.quote(img_url)
            filename = img_url_parsed.split('/')[-1]
            if sub in img_url and not os.path.exists(os.path.join(destination_dir, filename)):
                img_url = urllib.parse.urljoin(main, img_url)
                img_url = urllib.parse.quote(img_url, safe=':/')
                img_name = os.path.join(destination_dir, os.path.basename(img_url))
                req = urllib.request.Request(img_url, headers=headers)
                try:
                    with urllib.request.urlopen(req) as response:
                        with open(img_name, 'wb') as outfile:
                            outfile.write(response.read())
                    time.sleep(0.1)
                except urllib.error.URLError:
                    remember = i
                    i = 0
            i += 1
            if not only_check:
                if i > remember:
                    progress_bar.update(1)
        if not only_check:
            progress_bar.close()
    else:
        print(f"Failed to fetch {url}, try again later.")
        return None
    if not only_check:
        print('Widgets downloaded and updated')


_download_widgets()
os.makedirs('output-images', exist_ok=True)


def _size_identification(img_name, show_circles=False):
    """
    This function identifies the size of the widgets in the image. The function returns the size of the widgets in the
    image. If the show_circles parameter is set to True, the function will show the circles that were detected in the
    image.
    :param img_name: str
    :param show_circles: bool
    :return: size_widget: int
    """
    image_to_check = _screenshot_loading(img_name)
    if image_to_check is None:
        return None
    blurred = cv.GaussianBlur(image_to_check, (5, 5), 0)
    circles = cv.HoughCircles(blurred, cv.HOUGH_GRADIENT, dp=1, minDist=30,
                              param1=65, param2=55, minRadius=15, maxRadius=60)
    if circles is None:
        return None
    size_widget = (np.round(np.median(circles[0, :, 2])*2)+1).astype(dtype='int64')
    if show_circles:
        image_to_show = cv.cvtColor(image_to_check, cv.COLOR_GRAY2BGR)
        for circle in circles[0, :]:
            cv.circle(image_to_show, (np.round(circle[0]).astype(int), np.round(circle[1]).astype(int)),
                      np.round(circle[2]).astype(int), (0, 0, 255), 2)
        cv.imshow('circles', image_to_show)
        cv.waitKey(0)
        cv.destroyAllWindows()
        cv.waitKey(1)
    return size_widget


def _get_sizes(img_name):
    """
    This function calculates the final size of the widgets to be used in the image processing and the number of pixels to
    keep in the widgets
    :param img_name: str
    :return: final_size: int, pixels_to_keep: int
    """
    widget_size = 100
    target_size = _size_identification(img_name)
    if target_size is None:
        return None, None
    pixels_to_keep = 70
    ratio = target_size/widget_size
    final_size = np.floor(ratio*pixels_to_keep).astype(dtype='int64')
    return final_size, pixels_to_keep


def _widget_loading(img_names_tgt, img_name, save_sample=False):
    """
    This function loads the widgets to be used in the image processing. The output is a 3D array with the number of
    widgets in the first dimension and the final size of the widgets in the second and third dimensions
    :param img_names_tgt: np.array
    :param img_name: str
    :param save_sample: bool
    :return: check_img: np.array
    """
    final_size, pixels_to_keep = _get_sizes(img_name)
    if final_size is None:
        return None
    image_size = len(cv.imread(img_names_tgt[0]))
    check_img = np.zeros((len(img_names_tgt), final_size, final_size), dtype='uint8')
    for i in range(0, len(img_names_tgt)):
        check_img[i, :, :] = cv.resize(cv.imread(img_names_tgt[i], cv.IMREAD_GRAYSCALE)
                                       [round((image_size-pixels_to_keep)/2):round((image_size+pixels_to_keep)/2)-1,
                                        round((image_size-pixels_to_keep)/2):round((image_size+pixels_to_keep)/2)-1],
                                        (final_size, final_size), interpolation=4)
    if save_sample:
        cv.imwrite('output-images/sample_widget.png', check_img[0, :, :])
    else:
        return check_img


def _get_widget_description():
    """
    This function gets the description of the widgets from the Orange Data Mining website. The function then writes a
    yaml file with the description of the widgets
    """
    url = 'https://orangedatamining.com/widget-catalog/'
    reqs = requests.get(url)
    soup = BeautifulSoup(reqs.text, 'html.parser')
    urls = []
    keys = []
    for link in soup.find_all('a'):
        if 'widget-catalog/' in link.get('href') and '<img' in str(link):
            urls.append('https://orangedatamining.com' + link.get('href'))
            link_str = str(link)
            key = link_str.split('src="')[1].split('"')[0].split('/')[-1].split('.png')[0]
            key = key.split('-')[0] + '/' + '-'.join(key.split('-')[1:])
            keys.append(urllib.parse.unquote(key))
    descriptions = dict()
    progress_bar = tqdm(total=len(urls), desc="Progress")
    for j in range(len(urls)):
        progress_bar.update(1)
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36'}
        if urls[j].split('widget-catalog/')[-1] == '':
            continue
        try:
            req = urllib.request.Request(urls[j], headers=headers)
            html = urllib.request.urlopen(req).read()
            soup = BeautifulSoup(html, 'html.parser')
            for script in soup(['script', 'style']):
                script.extract()
            text = soup.get_text()
            short_description = text.split('\n')[1]
            inputs = ''
            if 'Inputs' in text:
                i = 2
                while True:
                    if text.split('Inputs')[1].split('\n')[i] == '' or 'OrangeFAQ' in text.split('Inputs')[1].split('\n')[i]:
                        break
                    inputs += '- ' + text.split('Inputs')[1].split('\n')[i] + '\n'
                    i += 1
            inputs = inputs[:-1]
            if inputs == '' or inputs == 'None':
                inputs = 'No inputs'
            outputs = ''
            if 'Outputs' in text:
                i = 2
                while True:
                    if text.split('Outputs')[1].split('\n')[i] == '' or 'OrangeFAQ' in text.split('Outputs')[1].split('\n')[i]:
                        break
                    outputs += '- ' + text.split('Outputs')[1].split('\n')[i] + '\n'
                    i += 1
                description_after = text.split(outputs.split('- ')[-1])[1]
            outputs = outputs[:-1]
            if outputs == '' or outputs == 'None':
                outputs = 'No outputs'
                description_after = text.split(inputs.split('- ')[-1])[1]
            if '\n' in description_after:
                description = description_after.split('\n')[1]
            descriptions[keys[j]] = {'description': description, 'inputs': inputs, 'outputs': outputs, 'short description': short_description}
        except urllib.error.URLError or urllib.error.HTTPError:
            progress_bar.write('url not found: ' + urls[j])
            continue
    progress_bar.close()
    with open('data/widget-info/widget-descriptions.yaml', 'w') as file:
        yaml.dump(descriptions, file)


def _screenshot_loading(img_name):
    """
    This function loads the screenshot to be used to identify the widgets and their links. It loads a color image and
    converts it to grayscale
    :param img_name:
    :return: image_to_check: np.array
    """
    reader = imageio.get_reader(img_name)
    for pict in reader:
        frame = cv.cvtColor(pict, cv.COLOR_RGB2BGR)
        image_to_check = cv.cvtColor(frame, cv.COLOR_BGR2GRAY)
        return image_to_check


def _is_there_widget_creation(img_name, value_thresh=0.8):
    """
    This function creates a matrix with the information of the widgets present in the image. The first column of the
    matrix indicates the widget presence, the second column indicates the location of the widget in the image, the third
    and fourth columns indicate the size of the image, and the fifth column indicates the value of the match. The default
    value of the threshold is 0.81, but it can be changed according to the needs of the user. A lower value will increase
    the number of widgets detected, but it will also increase the number of false positives. A higher value will
    decrease the number of widgets detected, but it will also increase the number of false negatives
    :param img_name: str
    :param value_thresh: float
    :return: is_there_widget: np.array
    """
    widget_size = _size_identification(img_name)
    if widget_size is None:
        return None
    img_names_tgt = _get_filenames('widgets')
    check_img = _widget_loading(img_names_tgt, img_name)
    is_there_widget = np.zeros((len(check_img), 5), dtype='int64')
    image_to_check = _screenshot_loading(img_name)
    for j in range(len(img_names_tgt)):
        res = cv.matchTemplate(image_to_check, check_img[j, :, :], cv.TM_CCOEFF_NORMED)
        if ('Data-Datasets' in img_names_tgt[j] or 'Data-File' in img_names_tgt[j] or 'Polynomial%20Regression' in
                img_names_tgt[j]) or 'Distance%20File' in img_names_tgt[j] or 'Impute' in img_names_tgt[j]:
            tmp = np.sum(res > 0.70)
        elif 'Periodogram' in img_names_tgt[j] or 'dictyExpress' in img_names_tgt[j] or 'Correlogram' in img_names_tgt[j]:
            tmp = np.sum(res > 0.90)
        elif 'Ontology' in img_names_tgt[j]:
            tmp = np.sum(res > 1)
        else:
            tmp = np.sum(res > value_thresh)
        form = res.shape
        if tmp:
            loc_y_to_check, loc_x_to_check = (-widget_size, -widget_size)
            is_far = 0
            ordered_res = res.argsort(axis=None)
            for k in range(tmp):
                raveled_loc = ordered_res[-(k+1), None]
                loc_y, loc_x = np.unravel_index(raveled_loc, form)
                if all(np.abs(loc_x - loc_x_to_check) > widget_size) or all(np.abs(loc_y - loc_y_to_check) > widget_size):
                    is_far = is_far + 1
                    loc_x_to_check = np.append(loc_x_to_check, loc_x)
                    loc_y_to_check = np.append(loc_y_to_check, loc_y)
                    if is_there_widget[j - is_far, 0] == 0:
                        is_there_widget[j - is_far, 0] = -1 * is_far
                        is_there_widget[j - is_far, 1] = np.array(raveled_loc)
                        is_there_widget[j - is_far, 2] = np.array(form)[0, None]
                        is_there_widget[j - is_far, 3] = np.array(form)[1, None]
                        is_there_widget[j - is_far, 4] = np.round(res[loc_y, loc_x]*1000)
                    else:
                        found = 0
                        while found == 0:
                            if is_there_widget[j - is_far, 0] == 0:
                                is_there_widget[j - is_far, 0] = -1 * is_far
                                is_there_widget[j - is_far, 1] = np.array(raveled_loc)
                                is_there_widget[j - is_far, 2] = np.array(form)[0, None]
                                is_there_widget[j - is_far, 3] = np.array(form)[1, None]
                                is_there_widget[j - is_far, 4] = np.round(res[loc_y, loc_x]*1000)
                                found = 1
                            else:
                                is_far = is_far + 1
    j = 0
    ind_to_check = np.where(is_there_widget[:, 0] != 0)[0]
    while j < len(ind_to_check):
        form = tuple([is_there_widget[ind_to_check[j], 2], is_there_widget[ind_to_check[j], 3]])
        before = 0
        loc_y, loc_x = np.unravel_index(is_there_widget[ind_to_check[j], 1].astype('int64'),
                                        tuple([is_there_widget[ind_to_check[j], 2], is_there_widget
                                        [ind_to_check[j], 3]]))
        loc_y_to_check, loc_x_to_check = np.unravel_index(is_there_widget[ind_to_check, 1], form)
        if np.any(np.logical_and(np.abs(loc_x - loc_x_to_check[np.arange(len(ind_to_check)) != j]) < widget_size,
                                 np.abs(loc_y - loc_y_to_check[np.arange(len(ind_to_check)) != j]) < widget_size)):
            which = np.array(np.where(np.logical_and(np.abs(loc_x - loc_x_to_check) < widget_size,
                                                     np.abs(loc_y - loc_y_to_check) < widget_size) == 1)[0])
            which_greatest = np.argmax(is_there_widget[ind_to_check[which], 4])
            is_there_widget[ind_to_check[which[np.arange(len(which)) != which_greatest]], :] = 0
            ind_to_check = np.where(is_there_widget[:, 0])[0]
            if which_greatest != 0:
                before = 1
        if before == 0:
            j = j+1
    if np.sum(is_there_widget[:, 0] != 0) == 0:
        return None
    return is_there_widget


def draw_locations(img_name, return_img=False):
    """
    This function shows the locations of the widgets in the image by highlighting them with a rectangle. The widgets are
    also labeled with their name. If the return_img parameter is set to True, the function will return the image with
    the widgets highlighted
    :param img_name: str
    :param return_img: bool
    :return: image: np.array
    """
    widget_size = _size_identification(img_name)
    if widget_size is None:
        print('There is no widget in the image')
        return None
    thickness = np.round(widget_size/2).astype(dtype='int64')
    final_size, _ = _get_sizes(img_name)
    is_there_widget = _is_there_widget_creation(img_name)
    if is_there_widget is None:
        print('There is no widget in the image')
        return None
    indexes = np.where(is_there_widget[:, 0] != 0)[0]
    form = (is_there_widget[indexes[0], 2], is_there_widget[indexes[0], 3])
    coord_y, coord_x = np.unravel_index(is_there_widget[indexes, 1], form) + np.floor(final_size/2).astype(dtype='int64')
    image = _screenshot_loading(img_name)
    image = cv.cvtColor(image, cv.COLOR_GRAY2BGR)
    adjusted_element_index = np.zeros_like(indexes)
    for j in range(len(indexes)):
        cv.rectangle(image, (coord_x[j]-thickness, coord_y[j]-thickness), (coord_x[j]+thickness, coord_y[j]+thickness),
                     (0, 0, 255), 2)
        adjusted_element_index[j] = indexes[j] - min(is_there_widget[indexes[j], 0], 0) - len(is_there_widget)
        label = urllib.parse.unquote(_get_filenames('widgets/')[adjusted_element_index[j]])
        if len(label.split('-')) == 2:
            label = label.split('-')[1]
        else:
            label = label.split('-')[1] + '-' + label.split('-')[2]
        label = label.split('.')[0]
        cv.putText(image, label, (coord_x[j]-thickness, coord_y[j]-thickness-10), cv.FONT_HERSHEY_SIMPLEX,
                   widget_size/150, (0, 0, 255), 1)
    if len(np.unique(adjusted_element_index)) == 1:
        print('There is no widget in the image')
        return None
    if not return_img:
        cv.imwrite('output-images/widget_locations.png', image)
        cv.imshow('widget locations highlighted in red', image)
        cv.waitKey(0)
        cv.destroyAllWindows()
        cv.waitKey(1)
    else:
        return image


def draw_links(img_name):
    """
    This function shows the links between the widgets in the image by drawing a line between the widgets. If the antennas
    are also shown in the image it means that the links are detected only through the connected components of the links,
    otherwise the links are detected through the circle intersection algorithm used on the link direction.
    :param img_name: str
    """
    widget_size = _size_identification(img_name)
    _, link_img = _link_detection(img_name)
    if link_img is None or widget_size is None:
        print('There are no links in the image')
        return None
    image = _screenshot_loading(img_name)
    image = cv.cvtColor(image, cv.COLOR_GRAY2BGR)
    image[link_img == 255] = (0, 0, 255)
    cv.imwrite('output-images/highlighted-links.png', image)
    cv.imshow('links highlighted in red', image)
    cv.waitKey(0)
    cv.destroyAllWindows()
    cv.waitKey(1)


def draw_links_and_locations(img_name):
    """
    This function shows the widget positions as well as the links between the widgets in the image.
    :param img_name: str
    """
    image = draw_locations(img_name, return_img=True)
    if image is None:
        print('There is no widget in the image')
        return None
    _, link_img = _link_detection(img_name)
    if link_img is not None:
        image[link_img == 255] = (0, 0, 255)
        cv.imwrite('output-images/links-widget-locations.png')
        cv.imshow('links and widget locations highlighted in red', image)
        cv.waitKey(0)
        cv.destroyAllWindows()
        cv.waitKey(1)


def _widgets_from_image(img_name, return_list=True):
    """
    This function returns the list of widgets present in the image
    :param img_name: str
    :param return_list: bool
    :return: widget_list: list of tuples
    """
    is_there_widget = _is_there_widget_creation(img_name)
    if is_there_widget is None:
        return None
    img_names_tgt = _get_filenames('widgets/')
    ind_present = np.where(is_there_widget[:, 0] != 0)[0].astype(dtype='int64')
    adjusted_element_index = np.zeros_like(ind_present)
    widget_list = list([])
    for j in range(len(ind_present)):
        adjusted_element_index[j] = ind_present[j] - min(is_there_widget[ind_present[j], 0], 0)
        adjusted_element_index[j] = np.where(adjusted_element_index[j] > len(is_there_widget) - 1, adjusted_element_index[j] - len(is_there_widget), adjusted_element_index[j])
        tmp = urllib.parse.unquote(img_names_tgt[adjusted_element_index[j]])
        _, module_name = os.path.split(tmp)
        if len(module_name.split('-')) == 2:
            module, name = module_name.split('-')
        else:
            module = module_name.split('-')[0]
            name = module_name.split('-')[1] + '-' + module_name.split('-')[2]
        name = name.split('.')[0]
        widget_list.append((module, name))
    if len(np.unique(widget_list)) == 2:
        return None
    if return_list:
        return widget_list
    else:
        return is_there_widget


def _find_circle_intersection(label_binary_image, center, radius_size, prev_direction, connect_type=8):
    """
    This function finds the intersection of a hand build circle with the white pixels of the binary image. The function
    returns the points of intersection, the direction of the points relative to the center, and the index of the point
    that best fits the previous direction
    :param label_binary_image: np.array
    :param center: tuple
    :param radius_size: int
    :param prev_direction: float
    :param connect_type: int
    :return: found_points: np.array, found_direction: np.array, best_fit_index: int
    """
    circle_image = np.zeros_like(label_binary_image)
    cv.circle(circle_image, center, radius_size, 255, 1)
    output = cv.bitwise_and(label_binary_image, circle_image)
    num_links, labeled_link_image = cv.connectedComponents(output, connectivity=connect_type)
    found_points = np.zeros((num_links - 1, 2)).astype(dtype='int64')
    for i in range(1, num_links):
        loc_y, loc_x = np.where(labeled_link_image == i)
        loc_x = np.mean(loc_x).astype(dtype='int64')
        loc_y = np.mean(loc_y).astype(dtype='int64')
        found_points[i-1, :] = [loc_x, loc_y]
    found_direction = np.arctan2(found_points[:, 1] - center[1], found_points[:, 0] - center[0]) * 180/np.pi
    found_direction = np.where(found_direction >= 0, found_direction, 360 + found_direction)
    best_fit_index = np.argmin(np.abs(found_direction - prev_direction))
    return found_points, found_direction, best_fit_index


def _link_detection(img_name, show_process=False, show_conn_comp=False):
    """
    This function detects the links between the widgets in the image. The function returns a matrix with the number of
    links between the widgets. The function also returns an image with the links highlighted. If the show_process
    parameter is set to True, the function will show the process of the link detection.
    :param img_name: str
    :param show_process: bool
    :return: links: np.array, link_img: np.array
    """
    is_there_widget = _widgets_from_image(img_name, False)
    if is_there_widget is None:
        return None, None
    non_adj_index = np.where(is_there_widget[:, 0] != 0)[0]
    adj_index = non_adj_index - np.where(is_there_widget[non_adj_index, 0] > 0, 0, is_there_widget[non_adj_index, 0])
    widgets_unique, widgets_num = np.unique(adj_index, return_counts=True)
    which_multiple = widgets_unique[np.where(widgets_num > 1)[0]]
    if which_multiple.size > 0:
        checked_ind = np.zeros((len(which_multiple), np.max(widgets_num)), dtype='int64')
    img_names_tgt = _get_filenames('widgets/')
    final_size, _ = _get_sizes(img_name)
    widget_size = _size_identification(img_name)
    tol = round(widget_size/2)
    # first dimension of the matrix will indicate what's the receiving widget and second dimension will be the widget
    # from which the link is coming
    links = np.zeros(([len(img_names_tgt), len(img_names_tgt)]), dtype='int64')
    # image processing to extract connected components
    tmp_img = _screenshot_loading(img_name)
    indexes = np.where(is_there_widget[:, 0] != 0)[0]
    form = (is_there_widget[indexes[0], 2], is_there_widget[indexes[0], 3])
    coord_y, coord_x = np.unravel_index(is_there_widget[indexes, 1], form) + np.floor(final_size/2).astype(dtype='int64')
    binary_image = np.where(tmp_img > 180, 0, 255).astype(np.uint8)
    if widget_size < 50:
        struct_elem = cv.getStructuringElement(cv.MORPH_CROSS, (3, 3))
        binary_image = cv.dilate(binary_image, struct_elem, iterations=1)
    # identifying connected components location and potential links
    num_lab, labels_im = cv.connectedComponents(binary_image)
    if show_conn_comp:
        im2show = np.zeros((binary_image.shape[0], binary_image.shape[1], 3), dtype=np.uint8)
        colors = [np.random.randint(0, 255, size=3).tolist() for i in range(num_lab)]
        for label in range(1, num_lab):  # label 0 is the background
            im2show[labels_im == label] = colors[label]
        cv.imshow('connected components', im2show)
        cv.waitKey(0)
        cv.destroyAllWindows()
        cv.waitKey(1)
        cv.imwrite('output-images/connected-components.png', im2show)
    labels_in = np.zeros((len(indexes), num_lab))
    labels_out = np.zeros((len(indexes), num_lab))
    iterate = len(coord_x)
    link_img = np.zeros_like(tmp_img)
    for j in range(iterate):
        tmp_in = np.unique(labels_im[(coord_y[j]-tol):(coord_y[j]+tol), (coord_x[j]-tol):(coord_x[j])])
        tmp_out = np.unique(labels_im[(coord_y[j]-tol):(coord_y[j]+tol), (coord_x[j]):(coord_x[j]+tol)])
        labels_in[j, 0:len(tmp_in)-1] = tmp_in[tmp_in != 0]
        labels_out[j, 0:len(tmp_out)-1] = tmp_out[tmp_out != 0]
    for j in range(iterate):
        in_to_check = labels_in[j, labels_in[j, :] != 0]
        for k in range(len(in_to_check)):
            value1 = 9
            value2 = 9
            except_element = np.arange(iterate) != j
            check_presence = labels_out[except_element, :] == in_to_check[k]
            which_present = np.any(check_presence, axis=1)
            check_presence = labels_in[except_element, :] == in_to_check[k]
            which_present = np.logical_or(which_present, np.any(check_presence, axis=1))
            if np.sum(which_present) == 1:
                link_img[labels_im == in_to_check[k]] = 255
                non_adj_index = indexes[j]
                if non_adj_index - min(is_there_widget[non_adj_index, 0], 0) in which_multiple:
                    which_which_multiple = np.where(which_multiple == non_adj_index - is_there_widget[non_adj_index, 0])[0]
                    if non_adj_index in checked_ind:
                        value2 = np.where(checked_ind[which_which_multiple, :] == non_adj_index)[1][0] + 1
                    else:
                        value2 = np.where(checked_ind[which_which_multiple, :] == 0)[1][0] + 1
                        checked_ind[which_which_multiple, value2 - 1] = non_adj_index
                in_element_index = non_adj_index - min(is_there_widget[non_adj_index, 0], 0) - len(is_there_widget)
                non_adj_index = indexes[except_element][which_present]
                if non_adj_index - min(is_there_widget[non_adj_index, 0], 0) in which_multiple:
                    which_which_multiple = np.where(which_multiple == non_adj_index - is_there_widget[non_adj_index, 0])[0]
                    if non_adj_index in checked_ind:
                        value1 = np.where(checked_ind[which_which_multiple, :] == non_adj_index)[1][0] + 1
                    else:
                        value1 = np.where(checked_ind[which_which_multiple, :] == 0)[1][0] + 1
                        checked_ind[which_which_multiple, value1 - 1] = non_adj_index
                which_indexes = non_adj_index - min(is_there_widget[non_adj_index, 0], 0) - len(is_there_widget)
                if value1 != 9 or value2 != 9:
                    if links[in_element_index, which_indexes] == 0:
                        links[in_element_index, which_indexes] = int(str(int(value1)) + str(int(value2)))
                    else:
                        links[in_element_index, which_indexes] = int(str(int(links[in_element_index, which_indexes])) +
                                                                     str(int(value1)) + str(int(value2)))
                else:
                    links[in_element_index, which_indexes] += 1
                break
            elif np.sum(which_present) > 1:
                label_binary_image = np.where(labels_im == in_to_check[k], 255, 0).astype(dtype='uint8')
                start_points, _, _ = _find_circle_intersection(label_binary_image, (coord_x[j]-round(widget_size/3.5),
                                                                                   coord_y[j]), round(widget_size/2), 0)
                actual_start = start_points[start_points[:, 0] < coord_x[j]-tol, :]
                non_adj_index = indexes[j]
                if non_adj_index - min(is_there_widget[non_adj_index, 0], 0) in which_multiple:
                    which_which_multiple = np.where(which_multiple == non_adj_index - is_there_widget[non_adj_index, 0])[0]
                    if non_adj_index in checked_ind:
                        value2 = np.where(checked_ind[which_which_multiple, :] == non_adj_index)[1][0] + 1
                    else:
                        value2 = np.where(checked_ind[which_which_multiple, :] == 0)[1][0] + 1
                        checked_ind[which_which_multiple, value2 - 1] = non_adj_index
                in_element_index = non_adj_index - min(is_there_widget[non_adj_index, 0], 0) - len(is_there_widget)
                for i in range(len(actual_start)):
                    prev_direction = 180
                    center = tuple(actual_start[i])
                    n_iter = 0
                    first_iter = True
                    tmp_link_img = np.zeros_like(tmp_img)
                    same_link = 0
                    while True:
                        found_points, found_direction, best_fit_index = _find_circle_intersection(label_binary_image, center,
                                                                                                 round(widget_size/10), prev_direction, 4)
                        if n_iter > 100 or same_link == 3:
                            break
                        if same_link < 3 and link_img[center[1], center[0]] == 255:
                            same_link += 1
                        elif same_link > 1:
                            same_link = 0
                        if abs(abs(found_direction[best_fit_index] - prev_direction) - 180) < 20:
                            center = (center[0]+1, center[1])
                        elif first_iter and abs(abs(found_direction[best_fit_index] - prev_direction) - 180) < 70:
                            center = (center[0]-1, center[1])
                        else:
                            cv.line(tmp_link_img, center, tuple(found_points[best_fit_index, :]), 255, 2)
                            center = tuple(found_points[best_fit_index, :])
                            prev_direction = found_direction[best_fit_index]
                        if show_process:
                            image_to_show = cv.cvtColor(tmp_img, cv.COLOR_GRAY2BGR)
                            cv.drawMarker(image_to_show, center, (0, 0, 255), cv.MARKER_CROSS, round(widget_size/6), 2)
                            cv.circle(image_to_show, center, round(widget_size/10), (0, 0, 255), 2)
                            cv.imshow('image', image_to_show)
                            cv.waitKey(100)
                        n_iter += 1
                        which_present = np.logical_and(abs(center[0] - coord_x[except_element]) - tol*1.5 < 0,
                                                       abs(center[1] - coord_y[except_element]) - tol*0.5 < 0)
                        if np.any(which_present):
                            non_adj_index = indexes[except_element][which_present]
                            if non_adj_index - min(is_there_widget[non_adj_index, 0], 0) in which_multiple:
                                which_which_multiple = np.where(which_multiple == non_adj_index - is_there_widget[non_adj_index, 0])[0]
                                if non_adj_index in checked_ind:
                                    value1 = np.where(checked_ind[which_which_multiple, :] == non_adj_index)[1][0] + 1
                                else:
                                    value1 = np.where(checked_ind[which_which_multiple, :] == 0)[1][0] + 1
                                    checked_ind[which_which_multiple, value1 - 1] = non_adj_index
                            out_element_index = non_adj_index - min(is_there_widget[non_adj_index, 0], 0) - len(is_there_widget)
                            if value1 != 9 or value2 != 9:
                                if links[in_element_index, out_element_index] == 0:
                                    links[in_element_index, out_element_index] = int(str(int(value1)) + str(int(value2)))
                                else:
                                    links[in_element_index, out_element_index] = int(str(int(links[in_element_index, out_element_index])) +
                                                                                     str(int(value1)) + str(int(value2)))
                            else:
                                links[in_element_index, out_element_index] += 1
                            link_img[tmp_link_img == 255] = 255
                            break
                        first_iter = False
                break
    if show_process:
        cv.destroyAllWindows()
        cv.waitKey(1)
    return links, link_img


def _extract_workflow_from_image(img_name, show_process=False):
    """
    This function returns the list of links present in the image
    :param img_name: str
    :param show_process: bool
    :return: link_list: Widget
    """
    links, _ = _link_detection(img_name, show_process)
    if links is None:
        return None
    img_names_tgt = _get_filenames('widgets/')
    if np.sum(links) == 0:
        return None
    a, b = np.where(links != 0)
    link_list = list([])
    for i in range(len(a)):
        if links[a[i], b[i]] == 1:
            tmp = urllib.parse.unquote(img_names_tgt[b[i]])
            _, module_name = os.path.split(tmp)
            if len(module_name.split('-')) == 2:
                module1, name1 = module_name.split('-')
            else:
                module1 = module_name.split('-')[0]
                name1 = module_name.split('-')[1] + '-' + module_name.split('-')[2]
            name1 = name1.split('.')[0]
            tmp = urllib.parse.unquote(img_names_tgt[a[i]])
            _, module_name = os.path.split(tmp)
            if len(module_name.split('-')) == 2:
                module2, name2 = module_name.split('-')
            else:
                module2 = module_name.split('-')[0]
                name2 = module_name.split('-')[1] + '-' + module_name.split('-')[2]
            name2 = name2.split('.')[0]
            for j in range(links[a[i], b[i]]):
                link_list.append(((module1, name1), (module2, name2)))
        else:
            info = str(links[a[i], b[i]])
            tmp = urllib.parse.unquote(img_names_tgt[b[i]])
            _, module_name = os.path.split(tmp)
            if len(module_name.split('-')) == 2:
                module1, name1 = module_name.split('-')
            else:
                module1 = module_name.split('-')[0]
                name1 = module_name.split('-')[1] + '-' + module_name.split('-')[2]
            name1 = name1.split('.')[0]
            tmp = urllib.parse.unquote(img_names_tgt[a[i]])
            _, module_name = os.path.split(tmp)
            if len(module_name.split('-')) == 2:
                module2, name2 = module_name.split('-')
            else:
                module2 = module_name.split('-')[0]
                name2 = module_name.split('-')[1] + '-' + module_name.split('-')[2]
            name2 = name2.split('.')[0]
            for j in range(int(len(str(links[a[i], b[i]]))/2)):
                info1 = info[2*j]
                info2 = info[2*j+1]
                if info1 == '9':
                    info1 = ''
                else:
                    info1 = ' #'+info1
                if info2 == '9':
                    info2 = ''
                else:
                    info2 = ' #'+info2
                link_list.append(((module1, name1 + info1), (module2, name2 + info2)))
    return link_list


class Widget:
    """
    This class represents a widget as a tuple of two strings.
    """
    def __init__(self, module, name=None):
        if name is None:
            if '/' in module:
                module, name = module.split('/')
            else:
                raise ValueError("the widget must be a tuple of two strings or a string with a '/' separator")
        self.module = module
        self.name = name

    def __repr__(self):
        """
        Custom representation for the Widget object.
        """
        return f"Widget({self.module}/{self.name})"

    def __str__(self):
        """
        Custom string representation for the Widget object.
        """
        return self.module + '/' + self.name

    def __eq__(self, other):
        """
        Custom equality comparison for the Widget object.
        :param other: any
        :return: bool
        """
        if isinstance(other, Widget):
            return self.module == other.module and self.name == other.name
        return False

    def __hash__(self):
        return hash((self.module, self.name))

    def get_description(self, short_description=False):
        """
        Returns the description of the widget.
        :param short_description: bool
        :return: description: str
        """
        try:
            with open('data/widget-info/widget-descriptions.yaml', 'r') as file:
                descriptions = yaml.full_load(file)
        except FileNotFoundError:
            _get_widget_description()
            with open('data/widget-info/widget-descriptions.yaml', 'r') as file:
                descriptions = yaml.full_load(file)
        key = self.module + '/' + self.name.split(' #')[0]
        if short_description:
            try:
                return [descriptions[key]['short description'], descriptions[key]['inputs'], descriptions[key]['outputs']]
            except KeyError:
                return None
        else:
            try:
                return [descriptions[key]['description'], descriptions[key]['inputs'], descriptions[key]['outputs']]
            except KeyError:
                return None


def _augment_widget_list(widget_list, present_widgets, goal=None, n=20, k=4):
    """
    This function augments the widget list given as input by adding widgets that are similar to the ones in the list. The
    similarity is assessed through the embedding of the widget descriptions performed by the BERT model and saved by the
    get_embedding function. n represents the number of widgets to have in the final list, and k represents the number of
    similar widgets to consider for each widget in the list. If the goal is not None, the function will return the widgets
    that are similar to the goal widget. Otherwise, the function will return the widgets that are similar to the widgets
    in the list.
    :param widget_list: list
    :param present_widgets: list
    :param goal: str
    :param n: int
    :param k: int
    :return: augmented_list: list
    """
    with open('data/widget-info/widget-embeddings.pkl', 'rb') as f:
        embeddings = pickle.load(f)
    if len(widget_list) >= n:
        if goal is None:
            return widget_list[:n]
        else:
            similarity_matrix = np.zeros((len(widget_list), 1))
            goal_embedding = _get_embedding('The goal of the widget is ' + goal)
            for i in range(len(widget_list)):
                similarity_matrix[i, 0] = cosine_similarity([embeddings[str(widget_list[i])]], [goal_embedding])
            similar_indices = similarity_matrix[:, 0].argsort(axis=0)[-n:][::-1]
            return [widget_list[i] for i in similar_indices]
    widget_names = list(embeddings.keys())
    present_widgets = present_widgets + widget_list
    similarity_matrix = np.zeros((len(widget_names), len(present_widgets)))
    if goal is not None:
        goal_embedding = _get_embedding('The goal of the widget is ' + goal)
        for i in range(len(widget_names)):
            if Widget(widget_names[i]) in present_widgets:
                continue
            similarity_matrix[i, 0] = cosine_similarity([embeddings[widget_names[i]]], [goal_embedding])
            k = 1
    else:
        for i in range(len(widget_names)):
            if Widget(widget_names[i]) in present_widgets:
                continue
            for j in range(len(present_widgets)):
                similarity_matrix[i, j] = cosine_similarity([embeddings[widget_names[i]]], [embeddings[str(present_widgets[j])]])
    for i in range(len(widget_names)):
        similarity_matrix[i, 0] = similarity_matrix[i, similarity_matrix[i, :].argsort()[-k:][::-1]].mean()
    similar_indices = similarity_matrix[:, 0].argsort(axis=0)[-(n-len(widget_list)):][::-1]
    addition = []
    for i in similar_indices:
        addition.append(Widget(widget_names[i]))
    augmented_list = widget_list + addition
    return augmented_list


def _get_embedding(text=None):
    """
    This function returns the embedding of the given text using the DistilBERT model. If the text is None, the function
    will return the embeddings of the widget descriptions and save them in a pickle file. Otherwise, the function will
    return the embedding of the given text.
    :param text: str
    :return: embeddings: np.array
    """
    model_name = "bert-base-uncased"
    tokenizer = BertTokenizer.from_pretrained(model_name)
    model = BertModel.from_pretrained(model_name)
    embeddings = dict({})
    if text is None:
        with open('data/widget-info/widget-descriptions.yaml', 'r') as file:
            widget_descriptions = yaml.safe_load(file)
        progress_bar = tqdm(total=len(widget_descriptions), desc='Progress')
        for widget, info in widget_descriptions.items():
            inputs = tokenizer(info['description'], return_tensors="pt", truncation=True, padding='max_length', max_length=512)
            with torch.no_grad():
                outputs = model(**inputs)
            embeddings.update({widget: outputs.last_hidden_state.mean(dim=1).squeeze().numpy()})
            progress_bar.update(1)
        progress_bar.close()
        with open('data/widget-info/widget-embeddings.pkl', 'wb') as f:
            pickle.dump(embeddings, f)
    elif isinstance(text, str):
        inputs = tokenizer(text, return_tensors="pt", truncation=True, padding='max_length', max_length=512)
        with torch.no_grad():
            outputs = model(**inputs)
        embeddings = outputs.last_hidden_state.mean(dim=1).squeeze().numpy()
        return embeddings


class Workflow:
    """
    This class represents a collection of workflows as a list of tuples of tuples.
    """
    def __init__(self, data, name=None):
        if isinstance(data, str):
            self.path = data
            name = data
            data = _extract_workflow_from_image(data)
        elif not isinstance(data, list):
            raise TypeError("the input must be a list of tuples of tuples or a string, got " + str(type(data)) + " instead")
        if name is None:
            self.path = 'unknown'
        else:
            self.path = name
        if data is None:
            if _size_identification(self.path) is not None:
                is_there_widget = _is_there_widget_creation(self.path)
                img_names_tgt = _get_filenames('widgets/')
                ind_present = np.where(is_there_widget[:, 0] != 0)[0].astype(dtype='int64')
                adjusted_element_index = ind_present - min(is_there_widget[ind_present, 0], 0)
                adjusted_element_index = np.where(adjusted_element_index > len(is_there_widget) - 1, adjusted_element_index - len(is_there_widget), adjusted_element_index)
                tmp = urllib.parse.unquote(img_names_tgt[adjusted_element_index[0]])
                _, module_name = os.path.split(tmp)
                if len(module_name.split('-')) == 2:
                    module, name = module_name.split('-')
                else:
                    module = module_name.split('-')[0]
                    name = module_name.split('-')[1] + '-' + module_name.split('-')[2]
                name = name.split('.')[0]
                data = [((module, name), ('None', 'None'))]
            else:
                raise ValueError("No links found in the image")
        link_list = []
        for i in data:
            if isinstance(i[0], Widget):
                first = (i[0].module, i[0].name)
            else:
                first = i[0]
            if isinstance(i[1], Widget):
                second = (i[1].module, i[1].name)
            else:
                second = i[1]
            link_list.append((first, second))
        self.data = link_list

    def __repr__(self):
        """
        Custom representation for the Workflow object.
        """
        return f"Workflow({self.path.split('/')[-1]}, {len(self.get_widgets())} widgets, {len(self.data)} links)"

    def __str__(self):
        """
        Custom string representation for the Workflow object.
        """
        return self._get_order()[0]

    def __len__(self):
        """
        Returns the number of links in the workflow.
        """
        return len(self.data)

    def __getitem__(self, item):
        """
        Returns the link at the given index.
        """
        return self.data[item]

    def _get_order(self):
        """
        Returns the order of the widgets in the workflow.
        :return: link_text: str, widget_text_list: list
        """
        links_present = []
        widget_in = []
        widget_out = []
        for widget_tuple in self.data:
            widget1 = widget_tuple[0][0] + '/' + widget_tuple[0][1]
            widget2 = widget_tuple[1][0] + '/' + widget_tuple[1][1]
            links_present.append(widget1 + ' -> ' + widget2)
            widget_in.append(widget1)
            widget_out.append(widget2)
        link_dict = dict()
        widget_in_unique = np.unique(widget_in)
        for i in range(len(widget_in_unique)):
            for j in range(len(widget_in)):
                if widget_in_unique[i] == widget_in[j]:
                    if widget_in_unique[i] not in link_dict:
                        link_dict[widget_in_unique[i]] = {widget_out[j]}
                    else:
                        link_dict[widget_in_unique[i]].add(widget_out[j])
        widget_order = np.flip(tuple(TopologicalSorter(link_dict).static_order()))
        link_text = ''
        widget_text_list = []
        dupl = {}
        for i in widget_order:
            if 'None' not in i:
                widget_text_list.append(i)
            for j in range(len(widget_in)):
                if i == widget_in[j]:
                    if '(' in widget_in[j]:
                        try:
                            if widget_in[j].split(' #')[1][0] not in dupl[widget_in[j].split(' #')[0]]:
                                instance_num = dupl[widget_in[j].split(' #')[0]]['how many instances'] + 1
                                dupl[widget_in[j].split(' #')[0]][widget_in[j].split('#')[1][0]] = instance_num
                                dupl[widget_in[j].split(' #')[0]]['how many instances'] = instance_num
                        except KeyError:
                            dupl[widget_in[j].split(' #')[0]] = {widget_in[j].split(' #')[1][0]: 1, 'how many instances': 1}
                        first_part = widget_in[j].split(' #')[0] + ' #' + str(dupl[widget_in[j].split(' #')[0]][widget_in[j].split(' #')[1][0]])
                    else:
                        first_part = widget_in[j]
                    if '(' in widget_out[j]:
                        try:
                            if widget_out[j].split(' #')[1][0] not in dupl[widget_out[j].split('#')[0]]:
                                instance_num = dupl[widget_out[j].split(' #')[0]]['how many instances'] + 1
                                dupl[widget_out[j].split(' #')[0]][widget_out[j].split(' #')[1][0]] = instance_num
                                dupl[widget_out[j].split(' #')[0]]['how many instances'] = instance_num
                        except KeyError:
                            dupl[widget_out[j].split(' #')[0]] = {widget_out[j].split('#')[1][0]: 1, 'how many instances': 1}
                        second_part = widget_out[j].split(' #')[0] + ' #' + str(dupl[widget_out[j].split(' #')[0]][widget_out[j].split(' #')[1][0]])
                    else:
                        second_part = widget_out[j]
                    link_text += '- ' + first_part + ' -> ' + second_part + '\n'
        return link_text, widget_text_list

    def get_widgets(self):
        """
        Returns a list of all unique widgets in the workflows.
        :return: widget_list: list
        """
        widget_list = []
        widget_text_list = self._get_order()[1]
        for i in widget_text_list:
            if i not in widget_list:
                widget_list.append(Widget(i.split('/')[0], i.split('/')[1].split(' #')[0]))
        return widget_list

    def get_context(self, short_description=False):
        """
        Returns the context of the workflow.
        :param short_description: bool
        :return: context: str
        """
        context = ''
        for widget in self.get_widgets():
            descr, inputs, outputs = widget.get_description(short_description=short_description)
            context += str(widget) + ':\n' + 'Description:\n' + descr + '\n' + 'Inputs: ' + inputs + '\n\n' + 'Outputs: ' + outputs + '\n\n\n'
        return context

    def remove_widget(self, widget=None):
        """
        Removes all links that contain the given widget.
        :param widget: Widget
        :return: widget: Widget
        """
        np.random.seed(11)
        if widget is None:
            if len(self.data) == 2:
                widget = self.get_widgets()[-1]
            else:
                widget = np.random.choice(np.array(self.get_widgets()[-2:]))
        tmp = [link for link in self.data if (widget.module, widget.name) not in link]
        if len(tmp) == 0:
            print('There is only one link in the workflow, if a widget was eliminated there would be no links left, so the function did not do anything')
            return None
        self.data = tmp
        return widget

    def get_name(self, model='gpt-3.5-turbo-0125', return_name=True):
        """
        Returns the name of the workflow.
        :param model: str
        :param return_name: bool
        :return: name: str
        """
        query = get_workflow_name_prompt(self, return_query=True)
        name = _get_response(query, model)
        if return_name:
            return name
        else:
            print(name)

    def get_description(self, model='gpt-3.5-turbo-0125', return_description=True, concise_description=True):
        """
        Returns the description of the workflow.
        :param model: str
        :param return_description: bool
        :param concise_description: bool
        :return: description: str
        """
        query = get_workflow_description_prompt(self, return_query=True, concise_description=concise_description)
        description = _get_response(query, model)
        if return_description:
            return description
        else:
            print(description)

    def get_new_widget(self, model='gpt-3.5-turbo-0125', return_widget=True, goal='Not specified'):
        """
        Returns three new widgets that should be added to the workflow and provides an explanation for each.
        :param model: str
        :param return_widget: bool
        :param goal: str
        :return: new_widgets: str
        """
        query = get_new_widget_prompt(self, goal=goal, return_query=True)
        new_widgets = _get_response(query, model)
        new_widgets = yaml.safe_load(new_widgets.split('yaml\n')[-1].split('`')[0])
        if return_widget:
            return new_widgets
        else:
            print(new_widgets)


def _get_response(query, model='gpt-3.5-turbo-0125'):
    """
    This function sends a query to the OpenAI API and returns the response.
    :param query: str
    :param model: str
    :return: response: str
    """
    api_key = os.getenv('OPENAI_API_KEY')
    if api_key is None:
        print('OpenAI API key not found.')
        return None
    client = OpenAI(api_key=api_key, organization='org-FvAFSFT8g0844DCWV1T2datD')
    response = client.chat.completions.create(model=model,
                                              messages=[
                                                  {"role": "system", "content": "You are ChatGPT, a large language model trained by OpenAI. "
                                                                                "Answer as concisely as possible.\nKnowledge cutoff: 2021-09-01"
                                                                                "\nCurrent date: {CurrentDate}"},
                                                  {'role': 'user', 'content': query},
                                              ],
                                              temperature=0.5,
                                              top_p=0.3)
    response = response.choices[0].message.content
    return response


def _update_image_list():
    """
    This function finds all the subdirectories inside the orange-lecture-notes-web/public/chapters directory and checks
    if the  information about the images inside the subdirectories are updated. The function creates a yaml file with
    the information about the images present in the notebooks. If an image file has not been updated for 30 days it will
    be updated, otherwise it will stay the same.
    """
    markdown_names = _get_filenames('orange-lecture-notes-web/public/chapters', 'index.md')
    today = date.today()
    yaml_direct = 'image-analysis-results'
    os.makedirs(yaml_direct, exist_ok=True)
    progress_bar = tqdm(total=len(markdown_names), desc="Progress")
    try:
        with open(yaml_direct+'/image-list.yaml', 'r') as file:
            images = yaml.full_load(file)
    except FileNotFoundError:
        images = dict({})
    for i in range(len(markdown_names)):
        direct = markdown_names[i].split('/')[3:-1]
        direct = '/'.join(direct)
        with open(markdown_names[i], 'r') as file:
            data = file.read()
        data = data.split('\n')
        for j in range(len(data)):
            if 'title:' in data[j]:
                data = data[j].split('title: ')[1]
                if '"' in data:
                    title = data.split('"')[1]
                    break
                elif "'" in data:
                    title = data.split("'")[1]
                    break
                else:
                    title = data
                    break
        img_names = _get_filenames('orange-lecture-notes-web/public/chapters/' + direct)
        for j in range(len(img_names)):
            img_names[j] = img_names[j].split('/')[-1]
        if len(img_names) > 0:
            try:
                prev_date = images[direct]['date']
                prev_date = date.fromisoformat(prev_date)
            except KeyError:
                prev_date = date(2000, 1, 1)
            if direct not in images or direct in images and today - prev_date > timedelta(days=30):
                images[direct] = {'document-title': str(title),
                                  'date': str(today),
                                  'images': img_names}
        progress_bar.update(1)
    with open(yaml_direct+'/image-list.yaml', 'w') as file:
        yaml.dump(images, file, sort_keys=False)
    progress_bar.close()


def _update_widget_list():
    """
    This function finds all the images inside the orange-lecture-notes-web/public/chapters directory and checks if the
    information about the widgets present in the images are updated. The function creates a yaml file with the
    information about the widgets present in the images. If an image file has not been updated for 30 days it will be
    updated, otherwise it will stay the same.
    """
    today = date.today()
    yaml_direct = 'image-analysis-results'
    os.makedirs(yaml_direct, exist_ok=True)
    try:
        with open(yaml_direct+'/image-widgets.yaml', 'r') as file:
            widgets = yaml.full_load(file)
    except FileNotFoundError:
        widgets = dict({})
    img_names_to_check = _get_filenames('orange-lecture-notes-web/public/chapters/')
    progress_bar = tqdm(total=len(img_names_to_check), desc="Progress")
    for i in range(len(img_names_to_check)):
        size = _size_identification(img_names_to_check[i])
        img_name = img_names_to_check[i].split('/')[-1]
        path = img_names_to_check[i].split('/')[3:-1]
        path = '/'.join(path)
        key = path + '---' + img_name
        if size is not None:
            try:
                prev_date = widgets[key]['date']
                prev_date = date.fromisoformat(prev_date)
            except KeyError:
                prev_date = date(2000, 1, 1)
            if key not in widgets or key in widgets and today - prev_date > timedelta(days=30):
                widget_list = _widgets_from_image(img_names_to_check[i])
                if widget_list is not None:
                    widgets[key] = {'path': path,
                                    'filename': img_name,
                                    'date': str(today),
                                    'widgets': widget_list}
                else:
                    widgets[key] = {'path': path,
                                    'filename': img_name,
                                    'date': str(today),
                                    'widgets': None}
        else:
            try:
                prev_date = widgets[key]['date']
                prev_date = date.fromisoformat(prev_date)
            except KeyError:
                prev_date = date(2000, 1, 1)
            if key not in widgets or key in widgets and today - prev_date > timedelta(days=30):
                widgets[key] = {'path': path,
                                'filename': img_name,
                                'date': str(today),
                                'widgets': None}
        progress_bar.update(1)
    with open(yaml_direct+'/image-widgets.yaml', 'w') as file:
        yaml.dump(widgets, file, sort_keys=False)
    progress_bar.close()


def _update_image_links():
    """
    This function finds all the images inside the orange-lecture-notes-web/public/chapters directory and checks if the
    information about the links between the widgets present in the images are updated. The function creates a yaml file
    with the information about the links between the widgets present in the images. If an image file has not been
    updated for 30 days it will be updated, otherwise it will stay the same.
    """
    today = date.today()
    yaml_direct = 'image-analysis-results'
    os.makedirs(yaml_direct, exist_ok=True)
    try:
        with open(yaml_direct+'/image-links.yaml', 'r') as file:
            links = yaml.full_load(file)
    except FileNotFoundError:
        links = dict({})
    img_names_to_check = _get_filenames('orange-lecture-notes-web/public/chapters/')
    progress_bar = tqdm(total=len(img_names_to_check), desc="Progress")
    for i in range(len(img_names_to_check)):
        size = _size_identification(img_names_to_check[i])
        img_name = img_names_to_check[i].split('/')[-1]
        path = img_names_to_check[i].split('/')[3:-1]
        path = '/'.join(path)
        key = path + '---' + img_name
        if size is not None and len(path) > 0:
            try:
                prev_date = links[key]['date']
                prev_date = date.fromisoformat(prev_date)
            except KeyError:
                prev_date = date(2000, 1, 1)
            if key not in links or key in links and today - prev_date > timedelta(days=30):
                link_list = _extract_workflow_from_image(img_names_to_check[i])
                if link_list is not None:
                    links[key] = {'path': path,
                                  'filename': img_name,
                                  'date': str(today),
                                  'links': link_list}
                else:
                    links[key] = {'path': path,
                                  'filename': img_name,
                                  'date': str(today),
                                  'links': None}
        elif len(path) > 0:
            try:
                prev_date = links[key]['date']
                prev_date = date.fromisoformat(prev_date)
            except KeyError:
                prev_date = date(2000, 1, 1)
            if key not in links or key in links and today - prev_date > timedelta(days=30):
                links[key] = {'path': path,
                              'filename': img_name,
                              'date': str(today),
                              'links': None}
        progress_bar.update(1)
    with open(yaml_direct+'/image-links.yaml', 'w') as file:
        yaml.dump(links, file, sort_keys=False)
    progress_bar.close()


def _crop_workflows(directory_to_check='orange-lecture-notes-web/public/chapters', img_name=None, no_yaml=False):
    """
    This function extracts the workflows from the image. Given the information about the position of the widgets
    contained in the image, the function crops the image to obtain a new image with only the workflow.
    :param directory_to_check: str
    :param img_name: str
    :param no_yaml: bool
    """
    if img_name is None:
        try:
            with open('image-analysis-results/reference/correct-image-widgets.yaml', 'r') as file:
                widgets = yaml.full_load(file)
        except FileNotFoundError:
            print('The image-widgets.yaml file is missing, please run the update_widget_list function first, set the '
                  'no_yaml parameter to True or specify the specific image name with the img_name parameter')
            return None
        list_to_check = list([])
        for name in widgets:
            if widgets[name]['widgets'] is not None:
                list_to_check.append(directory_to_check + widgets[name]['path'] + '/' + widgets[name]['filename'])
    elif no_yaml:
        list_to_check = _get_filenames(directory_to_check)
    else:
        list_to_check = [img_name]
    progress_bar = tqdm(total=len(list_to_check), desc="Progress")
    for i in range(len(list_to_check)):
        path = 'cropped-workflows/'+list_to_check[i].split('chapters/')[-1]
        if not os.path.exists(path):
            is_there_widget = _widgets_from_image(list_to_check[i], False)
            img = cv.imread(list_to_check[i])
            widget_size = _size_identification(list_to_check[i])
            which_present = np.where(is_there_widget[:, 0] != 0)[0]
            y_locs, x_locs = np.unravel_index(is_there_widget[which_present, 1], (is_there_widget[which_present[0], 2],
                                                                                  is_there_widget[which_present[0], 3]))
            y_min = max(round(np.min(y_locs)-widget_size/4), 0)
            x_min = max(round(np.min(x_locs)-widget_size/1.6), 0)
            y_max = min(round(np.max(y_locs) + widget_size/0.6), img.shape[0])
            x_max = min(round(np.max(x_locs) + widget_size/0.6), img.shape[1])
            img = img[y_min:y_max, x_min:x_max]
            os.makedirs(os.path.dirname(path), exist_ok=True)
            cv.imwrite(path, img)
        progress_bar.update(1)
    progress_bar.close()


def _workflow_to_code(workflow, return_labels=False, orange_dataset=True):
    """
    This function returns the code of the widgets present in the image. The code is a binary vector where 1 indicates the
    presence of the widget or link and 0 indicates the absence of the widget.  If the return_labels parameter is set to
    True, the function will return the labels of the widgets present in the image. If the orange_dataset parameter is set
    to True, the function will return the code of only the enriched widgets and links.
    :param workflow: Workflow
    :param return_labels: bool
    :param orange_dataset: bool
    :return: code: np.array, label_list: np.array
    """
    yaml_direct = 'image-analysis-results'
    widget_list = []
    for widget in workflow.get_widgets():
        widget_list.append(str(widget).split(' #')[0])
    if not orange_dataset:
        try:
            with open('data/widget-info/widget-descriptions.yaml', 'r') as file:
                descriptions = yaml.full_load(file)
        except FileNotFoundError:
            _get_widget_description()
            with open('data/widget-info/widget-descriptions.yaml', 'r') as file:
                descriptions = yaml.full_load(file)
        list_of_widget = list(descriptions.keys())
        list_of_widget.remove('Data/File')
        list_of_widget.remove('Data/Paint Data')
        list_of_widget.remove('Data/Datasets')
        code = []
        for i in range(len(list_of_widget)):
            widget = list_of_widget[i]
            if widget in widget_list:
                code.append(widget_list.count(widget))
            else:
                code.append(0)
        code.append(widget_list.count('Data/File')+widget_list.count('Data/Paint Data')+widget_list.count('Data/Datasets'))
        label_list = list_of_widget + ['Data/File']
    else:
        try:
            with open(yaml_direct+'/widgets-analysis.yaml', 'r') as file:
                widgets_enriched = yaml.full_load(file)
        except FileNotFoundError:
            print('There is no image-analysis-results/widgets-analysis.yaml file to read, please run the data_analysis '
                  'program first or download the file from the repository')
            return None
        widgets_enriched.remove('Data/File')
        widgets_enriched.remove('Data/Paint Data')
        widgets_enriched.remove('Data/Datasets')
        code = []
        for i in range(len(widgets_enriched)):
            widget = widgets_enriched[i]
            if widget in widget_list:
                code.append(1)
            else:
                code.append(0)
        if 'Data/File' in widget_list or 'Data/Paint Data' in widget_list or 'Data/Datasets' in widget_list:
            code.append(1)
        else:
            code.append(0)
        label_list = widgets_enriched + ['Data/File']

    link_to_memorize = []
    which = []
    try:
        with open(yaml_direct+'/reference/correct-image-links.yaml', 'r') as file:
            links = yaml.full_load(file)
    except FileNotFoundError:
        print('There is no yaml file to read, please run the update_widget_list function first')
        return None
    if not orange_dataset:
        list_of_links = []
        for key in links:
            if links[key]['links'] is not None:
                w = Workflow(links[key]['links'])
                for link in w:
                    if '#' in link[0][1]:
                        link = ((link[0][0], link[0][1].split(' #')[0]), link[1])
                    if '#' in link[1][1]:
                        link = (link[0], (link[1][0], link[1][1].split(' #')[0]))
                    link_str = link[0][0] + '/' + link[0][1] + ' -> ' + link[1][0] + '/' + link[1][1]
                    if link_str not in list_of_links:
                        if 'Data/File' in link_str or 'Data/Paint Data' in link_str or 'Data/Datasets' in link_str:
                            link_to_memorize.append(link_str.split('Data/File')[-1].split('Data/Paint Data')[-1].split('Data/Datasets')[-1])
                        else:
                            list_of_links.append(link_str)
        link_to_memorize = np.unique(link_to_memorize)
        for i in list_of_links:
            if i in str(workflow):
                code.append(1)
            else:
                code.append(0)
        for i in ['Data/File', 'Data/Paint Data', 'Data/Datasets']:
            for j in range(len(link_to_memorize)):
                if i + link_to_memorize[j] in str(workflow):
                    which.append(j)
        which = np.unique(which)
        for i in range(len(link_to_memorize)):
            list_of_links.append('Data/File' + link_to_memorize[i])
            if i in which:
                code.append(1)
            else:
                code.append(0)
        label_list = np.append(label_list, list_of_links)
    else:
        try:
            with open(yaml_direct+'/links-analysis.yaml', 'r') as file:
                links_enriched = yaml.full_load(file)
        except FileNotFoundError:
            print('There is no image-analysis-results/links-analysis.yaml file to read, please run the data_analysis '
                  'program first or download the file from the repository')
            return None
        link_list = []
        for i in links_enriched:
            if 'Data/File' in i or 'Data/Paint Data' in i or 'Data/Datasets' in i:
                links_enriched.remove(i)
                link_to_memorize.append(i.split('Data/File')[-1].split('Data/Paint Data')[-1].split('Data/Datasets')[-1])
        for link in workflow.data:
            if '#' in link[0][1]:
                link = ((link[0][0], link[0][1].split(' #')[0]), link[1])
            if '#' in link[1][1]:
                link = (link[0], (link[1][0], link[1][1].split(' #')[0]))
            link = str(link)
            link_list.append(link)
        for i in links_enriched:
            if i in link_list:
                code.append(1)
            else:
                code.append(0)
        for i in ['Data/File', 'Data/Paint Data', 'Data/Datasets']:
            for j in range(len(link_to_memorize)):
                if i + link_to_memorize[j] in str(workflow):
                    which.append(j)
        which = np.unique(which)
        for i in range(len(link_to_memorize)):
            links_enriched.append('Data/File' + link_to_memorize[i])
            if i in which:
                code.append(1)
            else:
                code.append(0)
        label_list = np.append(label_list, links_enriched)

    if not return_labels:
        return code
    else:
        return label_list


def _create_dataset(orange_dataset=True):
    """
    This function creates an Excel file with the information about the widgets present in the images. The Excel file
    contains the name of the workflow, the path of the image, and the widgets present in the image. If the orange_dataset
    parameter is set to True, the function will create a dataset with only the enriched widgets and links.
    :param orange_dataset: bool
    """
    try:
        img_names_to_check = _get_filenames('cropped-workflows/')
    except FileNotFoundError:
        print('There are no cropped workflows, run the function crop_workflows() first')
        return None
    try:
        with open('image-analysis-results/reference/correct-image-links.yaml', 'r') as file:
            links = yaml.full_load(file)
    except FileNotFoundError:
        print('There is no information about the links, please run the update_image_links function first or '
              'download it from the repository')
        return None
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    progress_bar = tqdm(total=len(img_names_to_check), desc='Progress')
    sheet['A1'] = 'Workflow name'
    sheet['B1'] = 'Path'
    sheet['C1'] = 'Parent Folder'
    sheet['D1'] = 'Parent Subfolder'
    key = os.path.dirname(img_names_to_check[0].split('cropped-workflows/')[-1]) + '---' + img_names_to_check[0].split('/')[-1]
    labels = _workflow_to_code(Workflow(links[key]['links']), return_labels=True, orange_dataset=orange_dataset)
    for j in range(len(labels)):
        sheet.cell(row=1, column=j+5, value=labels[j])
    for i in range(len(img_names_to_check)):
        sheet['A'+str(i+2)] = img_names_to_check[i].split('/')[-1].replace('.png', '')
        sheet['B'+str(i+2)] = os.getcwd() + '/' + img_names_to_check[i]
        if os.path.dirname(img_names_to_check[i]).split('/')[-2] != 'cropped-workflows':
            sheet['C'+str(i+2)] = os.path.dirname(img_names_to_check[i]).split('/')[-2]
        else:
            sheet['C'+str(i+2)] = os.path.dirname(img_names_to_check[i]).split('/')[-1]
        sheet['D'+str(i+2)] = os.path.dirname(img_names_to_check[i]).split('/')[-1]
        key = os.path.dirname(img_names_to_check[i].split('cropped-workflows/')[-1]) + '---' + img_names_to_check[i].split('/')[-1]
        workflow = Workflow(links[key]['links'])
        code = _workflow_to_code(workflow, orange_dataset=orange_dataset)
        for j in range(len(code)):
            sheet.cell(row=i+2, column=j+5, value=int(code[j]))
        progress_bar.update(1)
    progress_bar.close()
    if orange_dataset:
        workbook.save('image-analysis-results/orange-dataset.xlsx')
    else:
        workbook.save('image-analysis-results/workflows-dataset.xlsx')


def _get_example_workflows(concise_description):
    """
    This function loads the names, descriptions and Workflows of 5 example workflows. If concise_description is set to
    True, the function will return a concise description of the workflow alongside the remaining information, otherwise
    a detailed description will be provided.
    :param concise_description: str
    :return: output_list: list of lists of str, list of tuples of tuples, str
    """
    folders = _get_filenames('data/workflows/samples')
    output_list = []
    for i in folders:
        image_name = i.split('/')[-1]
        with open('data/workflows/samples/samples.yaml', 'r') as file:
            info = yaml.full_load(file)
        workflow = info[image_name]['workflow']
        if concise_description:
            description = info[image_name]['concise']
        else:
            description = info[image_name]['detailed']
        output_list.append([image_name, workflow, description])
    return output_list


def find_similar_workflows(workflow, return_workflows=True, k=10, dist_type='euclidean adjusted', only_widgets=True, remove_widget=False):
    """
    This function loads the dataset from the Excel file and finds the k-closest workflows to the input workflow thanks
    to Euclidean distance between the code of the workflows
    :param workflow: Workflow
    :param remove_widget: bool or str
    :param return_workflows: bool
    :param k: int
    :param dist_type: str
    :param only_widgets: bool
    :param remove_widget: bool
    :return: closest_workflows: list
    """
    label_list = list(_workflow_to_code(workflow, return_labels=True, orange_dataset=False))
    for i in range(len(label_list)):
        if '->' in label_list[i]:
            how_many_widgets = i
            break
    if not only_widgets:
        how_many_widgets = len(label_list)
    if remove_widget:
        if isinstance(remove_widget, bool):
            removed_widget = workflow.remove_widget()
        elif isinstance(remove_widget, Widget):
            removed_widget = workflow.remove_widget(remove_widget)
        else:
            print('The remove_widget parameter must be a boolean or a Widget object')
            return None, None
    else:
        removed_widget = None
    workflow_code = np.array(_workflow_to_code(workflow, orange_dataset=False))[:how_many_widgets].astype(int)
    df = pd.read_excel('image-analysis-results/workflows-dataset.xlsx')
    code = df.iloc[:, 4:].values
    unique_code, idx = np.unique(code[:, :how_many_widgets], axis=0, return_index=True)
    if dist_type == 'euclidean adjusted':
        difference = np.where(unique_code - workflow_code < 0, 1, 0)
        difference[np.where(unique_code - workflow_code > 0)[0], np.where(unique_code - workflow_code > 0)[1]] = 0.2
        distances = difference.sum(axis=1)
    elif dist_type == 'euclidean':
        distances = np.linalg.norm(unique_code - workflow_code, axis=1)
    elif dist_type == 'cosine':
        distances = np.zeros(unique_code.shape[0])
        for i in range(unique_code.shape[0]):
            distances[i] = cosine_similarity((unique_code[i, :]).reshape(1, -1), workflow_code.reshape(1, -1))
        distances[np.where(distances == 1)] = 0
    else:
        print('The distance type must be either euclidean, euclidean adjusted or cosine')
        return None
    if dist_type == 'cosine':
        closest_idx = np.argsort(distances)[::-1][:k]
    else:
        closest_idx = np.argsort(distances)[:k]
    closest_workflows = []
    possible_widgets = []
    for i in closest_idx:
        widget_index = np.where(unique_code[i, :] - workflow_code > 0)[0]
        closest_workflows.append(df.iloc[idx[i], 1])
        if len(widget_index) == 0:
            continue
        for j in range(len(widget_index)):
            widget = label_list[widget_index[j]]
            if '>' in widget:
                break
            if widget not in possible_widgets:
                possible_widgets.append(widget)
    if return_workflows:
        similar_workflows = []
        with open('image-analysis-results/reference/correct-image-links.yaml', 'r') as file:
            links = yaml.full_load(file)
        for i in closest_workflows:
            filename = os.path.dirname(i.split('cropped-workflows/')[1]) + '---' + i.split('/')[-1]
            similar_workflows.append(Workflow(links[filename]['links'], i))
        return similar_workflows
    for i in range(len(possible_widgets)):
        possible_widgets[i] = Widget(possible_widgets[i].split('/')[0], possible_widgets[i].split('/')[1])
    return possible_widgets, removed_widget


def get_workflow_name_prompt(img_name, return_query=False):
    """
    This function uses the OpenAI API to generate a prompt that can be used to get the name of the workflow present in
    the image.
    :param img_name: str or Workflow
    :param return_query: bool
    :return: query: str
    """
    with open('data/prompts/prompt-intro.md', 'r') as file:
        query = file.read()
    with open('data/prompts/new-name-prompt.md', 'r') as file:
        query += file.read()
    examples = _get_example_workflows(concise_description=True)
    for example in examples:
        workflow = Workflow(example[1])
        query += '## Workflow:\nLinks in the workflow:\n' + str(workflow) + '\n\nWidget descriptions:\n'
        query += workflow.get_context(True)
        query += '## Name:\n' + example[0] + '\n-------------------------------\n\n'
    try:
        workflow = Workflow(img_name)
    except ValueError:
        print('No links found in the image')
        return None
    except TypeError:
        if isinstance(img_name, Workflow):
            workflow = img_name
        else:
            print('The img_name parameter must be a string or a Workflow object')
            return None
    query += '## Workflow:\nLinks in the workflow:\n' + str(workflow) + '\n\nWidget descriptions:\n'
    query += workflow.get_context(True)
    query += '## Name:\n'
    if return_query:
        return query
    else:
        print(query + '\n')


def get_workflow_description_prompt(img_name, return_query=False, concise_description=True):
    """
    This function uses the OpenAI API to generate a description of the workflow present in the image.
    :param img_name: str or Workflow
    :param return_query: bool
    :param concise_description: bool
    :return: query: str
    """
    with open('data/prompts/prompt-intro.md', 'r') as file:
        query = file.read()
    with open('data/prompts/new-description-prompt.md', 'r') as file:
        query += file.read()
    examples = _get_example_workflows(concise_description)
    for example in examples:
        workflow = Workflow(example[1])
        query += '## Workflow:\nLinks in the workflow:\n' + str(workflow) + '\n\nWidget descriptions:\n'
        widgets = workflow.get_widgets()
        for widget in widgets:
            descr, inputs, outputs = widget.get_description()
            query += str(widget) + ':\n' + 'Description:\n' + descr + '\n' + 'Inputs:\n' + inputs + '\n\n' + 'Outputs:\n' + outputs + '\n\n\n'
        query += '## Description:\n' + example[2] + '\n\n'
    try:
        workflow = Workflow(img_name)
    except ValueError:
        print('No links found in the image')
        return None
    except TypeError:
        if isinstance(img_name, Workflow):
            workflow = img_name
        else:
            print('The img_name parameter must be a string or a Workflow object')
            return None
    query += '## Workflow:\nLinks in the workflow:\n' + str(workflow) + '\n\nWidget descriptions:\n'
    for widget in workflow.get_widgets():
        descr, inputs, outputs = widget.get_description()
        query += str(widget) + ':\n' + 'Description:\n' + descr + '\n' + 'Inputs: ' + inputs + '\n\n' + 'Outputs: ' + outputs + '\n\n\n'
    query += '## Description:\n'
    if return_query:
        return query
    else:
        print(query + '\n')


def get_new_widget_prompt(img_name, goal='Not specified', remove_widget=False, return_query=False):
    """
    This function describes what widget can come next in the given workflow. The function uses the OpenAI API to provide
    which widgets can come next in the workflow and the reason for each widget. If the remove_widget parameter is set to
    True, the function will remove a random widget from the workflow and ask GPT to provide a widget on the newly
    obtained workflow. If the multiple parameter is set to True, the function will ask GPT to provide multiple widgets
    that can come next in the workflow as well as a reason for each widget.
    :param img_name: str, tuples of tuples of str or Workflow
    :param goal: str
    :param remove_widget: bool or Widget
    :param return_query: bool
    :return: query: str
    """
    with open('data/prompts/prompt-intro.md', 'r') as file:
        query = file.read()
    try:
        workflow = Workflow(img_name)
    except ValueError:
        print('No links found in the image')
        return None
    except TypeError:
        if isinstance(img_name, Workflow):
            workflow = img_name
        else:
            print('The img_name parameter must be a string or a Workflow object')
            return None
    possible_widgets, removed_widget = find_similar_workflows(workflow, remove_widget=remove_widget, return_workflows=False)
    if goal == 'Not specified':
        possible_widgets = _augment_widget_list(possible_widgets, present_widgets=workflow.get_widgets())
    else:
        possible_widgets = _augment_widget_list(possible_widgets, present_widgets=workflow.get_widgets(), goal=goal)
    query += '## Workflow:\nLinks in the workflow:\n' + str(workflow) + '\n\nWidget descriptions:\n'
    query += workflow.get_context(True)
    query += '## Possible widgets:\n'
    for widget in possible_widgets:
        descr, inputs, outputs = widget.get_description(True)
        query += str(widget) + ':\n' + 'Description:\n' + descr + '\n' + 'Inputs:\n' + inputs + '\n\n' + 'Outputs:\n' + outputs + '\n\n\n'
    query += '## Goal of the new widget:\n' + goal + '.\n\n'
    with open('data/prompts/new-widget-prompt.md', 'r') as file:
        query += file.read()
    if return_query:
        return query
    else:
        print(query + '\n')


def _workflow_detection_evaluation():
    """
    This function evaluates the performance of the workflow detection by comparing the detected workflows, widgets and
    links with the actual workflows, widgets and links.
    """
    with open('image-analysis-results/image-links.yaml', 'r') as file:
        links = yaml.full_load(file)
    with open('image-analysis-results/image-widgets.yaml', 'r') as file:
        widgets = yaml.full_load(file)
    with open('image-analysis-results/reference/correct-image-links.yaml') as file:
        actual_links = yaml.full_load(file)
    with open('image-analysis-results/reference/correct-image-widgets.yaml') as file:
        actual_widgets = yaml.full_load(file)

    correct_screenshots = 0
    correct_workflows = 0
    n_workflows = 0
    n_screenshots = 0
    correct_widgets = 0
    extra_widgets = 0
    n_widgets = 0
    correct_links = 0
    extra_links = 0
    n_links = 0
    for key in actual_links:
        correct = True
        n_screenshots += 1
        if actual_widgets[key]['widgets'] is None and widgets[key]['widgets'] is None:
            correct_screenshots += 1
            continue
        elif actual_widgets[key]['widgets'] is None:
            continue
        elif widgets[key]['widgets'] is None:
            n_workflows += 1
            continue
        n_workflows += 1
        for widget in actual_widgets[key]['widgets']:
            n_widgets += 1
            if widget in widgets[key]['widgets']:
                correct_widgets += 1
            else:
                correct = False
        for widget in widgets[key]['widgets']:
            if widget not in actual_widgets[key]['widgets']:
                extra_widgets += 1
                correct = False
        for link in actual_links[key]['links']:
            found = False
            n_links += 1
            if link in links[key]['links'] and '#' not in link[0][1] and '#' not in link[1][1]:
                correct_links += 1
            elif '#' in link[0][1] or '#' in link[1][1]:
                link = link[0][0] + '/'+ link[0][1].split(' #')[0] + ' -> ' + link[0][0] + '/'+ link[0][1].split(' #')[0]
                for link_to_check in links[key]['links']:
                    link_to_check = link_to_check[0][0] + '/'+ link_to_check[0][1].split(' #')[0] + ' -> ' + link_to_check[0][0] + '/'+ link_to_check[0][1].split(' #')[0]
                    if link in link_to_check:
                        correct_links += 1
                        found = True
                        break
                if not found:
                    correct = False
            else:
                correct = False
        for link in links[key]['links']:
            found_extra = False
            if link not in actual_links[key]['links'] and '#' not in link[0][1] and '#' not in link[1][1]:
                extra_links += 1
                correct = False
            elif '#' in link[0][1] or '#' in link[1][1]:
                link = link[0][0] + '/' + link[0][1].split(' #')[0] + ' -> ' + link[0][0] + '/' + link[0][1].split(' #')[0]
                for link_to_check in actual_links[key]['links']:
                    link_to_check = link_to_check[0][0] + '/' + link_to_check[0][1].split(' #')[0] + ' -> ' + link_to_check[0][0] + '/' + link_to_check[0][1].split(' #')[0]
                    if link in link_to_check:
                        found_extra = True
                        break
                if not found_extra:
                    extra_links += 1
                    correct = False
        if correct:
            correct_screenshots += 1
            correct_workflows += 1

    print('The workflow detection got ' + str(correct_screenshots/n_screenshots) + '% of screenshots correct')
    print('The workflow detection got ' + str(correct_workflows/n_workflows) + '% of workflows correct')
    print('The workflow detection got ' + str(correct_widgets/n_widgets) + '% of widgets correct with ' +
          str(extra_widgets) + ' extra widgets detected')
    print('The workflow detection got ' + str(correct_links/n_links) + '% of links correct with ' + str(extra_links) +
          ' extra links detected')


def _name_evaluation(model='gpt-3.5-turbo-0125'):
    """
    This function evaluates the performance of the Workflow.get_name function by comparing the output of the function with
    the reference name of the workflow, by comparing the two names using ChatGPT.
    """
    filenames = _get_filenames('data/workflows/evaluation/name-and-description')
    with open('data/prompts/text-comparison-prompt.md', 'r') as file:
        query_start = file.read()
    results = 0
    for name in filenames:
        workflow = Workflow(name)
        name_generated = workflow.get_name(model).replace('-', ' ')
        name_actual = name.split('/')[-1].replace('.png', '').replace('-', ' ')
        print('Evaluating the get_name function for the workflow: ' + name)
        query = query_start
        query += '\n\nText 1:\n' + name_generated + '\n\nText 2:\n' + name_actual + '\n\nScore:'
        response = _get_response(query, 'gpt-3.5-turbo-0125')
        score = [int(i) for i in response if i.isdigit()]
        if len(score) == 2:
            score = 10
        else:
            score = score[0]
        print('The score for the get_name function is ' + str(score) + ' out of 10\n\n')
        results += score
    print('The get_name function got a score of ' + str(results) + '%')


def _description_evaluation(model='gpt-3.5-turbo-0125', concise_description=True):
    """
    This function evaluates the performance of the Workflow.get_description function by comparing the output of the
    function with the actual description of the widget, by comparing the two descriptions using ChatGPT.
    :param model: str
    :param concise_description: bool
    """
    filenames = _get_filenames('data/workflows/evaluation/name-and-description')
    with open('data/prompts/text-comparison-prompt.md', 'r') as file:
        query_start = file.read()
    results = 0
    with open('data/workflows/evaluation/name-and-description/description-evaluation.yaml', 'r') as file:
        widgets_info = yaml.safe_load(file)
    for name in filenames:
        workflow = Workflow(name)
        description_generated = workflow.get_description(model, concise_description=concise_description)
        if concise_description:
            description_actual = widgets_info[name.split('/')[-1]]['concise']
        else:
            description_actual = widgets_info[name.split('/')[-1]]['detailed']
        print('Evaluating the get_description function for the workflow: ' + name)
        query = query_start
        query += '\n\nText 1:\n' + description_generated + '\n\nText 2:\n' + description_actual + '\n\nScore:'
        response = _get_response(query, 'gpt-3.5-turbo-0125')
        score = [int(i) for i in response if i.isdigit()]
        if len(score) == 2:
            score = 10
        else:
            score = score[0]
        print('The score for the get_description function is ' + str(score) + ' out of 10\n\n')
        results += score
    print('The get_description function got a score of ' + str(results) + '%')


def _new_widget_evaluation(check_response=True, model='gpt-3.5-turbo-0125', dist_type='euclidean adjusted', only_widgets=True):
    """
    This function evaluates the performance of the Workflow.get_new_widget function by comparing the output of the
    function with the actual widget that comes next in the workflow.
    check_response: bool
    model: str
    dist_type: str
    only_widgets: bool
    """
    n_correct = 0
    ignored = 0
    count = 0
    not_present = 0
    filenames = _get_filenames('data/workflows/evaluation/new-widgets')
    with open('data/workflows/evaluation/new-widgets/new-widget-evaluation.yaml', 'r') as file:
        workflows_info = yaml.safe_load(file)
    for name in filenames:
        try:
            workflow = Workflow(name)
        except ValueError:
            ignored += 1
            continue
        possible_widgets, _ = find_similar_workflows(workflow, return_workflows=False, dist_type=dist_type, only_widgets=only_widgets)
        if isinstance(workflows_info[name.split('/')[-1]]['goal'], list):
            for goal in workflows_info[name.split('/')[-1]]['goal']:
                aug_list = _augment_widget_list(possible_widgets, present_widgets=workflow.get_widgets(), goal=goal)
                print('Evaluating the new_widget_prompt function for the workflow: ' + name + ' for goal ' + goal)
                target_widget = workflows_info[name.split('/')[-1]]['widget'][goal]
                if check_response:
                    response = list(workflow.get_new_widget(goal=goal, model=model).keys())
                else:
                    response = ''
                count += 1
                if isinstance(target_widget, list):
                    found = 0
                    present = False
                    for i in target_widget:
                        if Widget(i) in aug_list:
                            present = present or True
                        if i in response:
                            found += 1
                            n_correct += 1
                            break
                    if found == 0:
                        print('The response does not contain the removed widget for the workflow: ' + name + ' for goal ' + goal)
                        if not present:
                            print('The widget is not in the possible widgets')
                            not_present += 1
                            if not check_response:
                                for i in aug_list:
                                    print(str(i))
                else:
                    if target_widget in response:
                        n_correct += 1
                    else:
                        print('The response does not contain the removed widget for the workflow: ' + name + ' for goal ' + goal)
                        if Widget(target_widget) not in aug_list:
                            print('The widget is not in the possible widgets')
                            not_present += 1
                            if not check_response:
                                for i in aug_list:
                                    print(str(i))
                print()
        else:
            print('Evaluating the new_widget_prompt function for the workflow: ' + name)
            aug_list = _augment_widget_list(possible_widgets, present_widgets=workflow.get_widgets(), goal=workflows_info[name.split('/')[-1]]['goal'])
            target_widget = workflows_info[name.split('/')[-1]]['widget']
            if check_response:
                response = list(workflow.get_new_widget(goal=workflows_info[name.split('/')[-1]]['goal'], model=model).keys())
            else:
                response = ''
            count += 1
            if isinstance(target_widget, list):
                found = 0
                present = False
                for i in target_widget:
                    if Widget(i) in aug_list:
                        present = present or True
                    if i in response:
                        found += 1
                        n_correct += 1
                        break
                if found == 0:
                    print('The response does not contain the removed widget for the workflow: ' + name)
                    if not present:
                        print('The widget is not in the possible widgets')
                        not_present += 1
                        if not check_response:
                            for i in aug_list:
                                print(str(i))
                print('\n')
            else:
                if target_widget not in response:
                    print('The response does not contain the removed widget for the workflow: ' + name)
                    if Widget(target_widget) not in aug_list:
                        print('The widget is not in the possible widgets')
                        not_present += 1
                        if not check_response:
                            for i in aug_list:
                                print(str(i))
                else:
                    n_correct += 1
                print('\n')
    print('The new_widget_prompt function predicts ' + str(n_correct) + ' out of ' + str(count) + ' workflows correctly')
    print('The accuracy of the new_widget_prompt function is ' + str(n_correct/count*100)[:4] + '%')
    print('The number of workflows that do not have the target widget suggested is ' + str(not_present))

#%%
